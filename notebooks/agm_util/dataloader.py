import pandas as pd
from glob import glob
from os import path
import numpy as np
from scipy.spatial.transform import Rotation as R
import openpyxl
import random
import pickle
from ipywidgets import FloatProgress


def ExcelCol2Index(s):
    ''' Utility function that takes the string representation of a column (like from Excel) and returns
        the index of the column.'''
    s = reversed(s.lower())
    ret = -1
    for i, v in enumerate(s):
        ret += (ord(v) - ord('a') + 1) * (26 ** i)
    return ret

def GetCol(sheet, s, r):
    ''' returns a List of the values in sheet, column s, over row indices r.'''
    #return [sheet.cell(x,ExcelCol2Index(s)).value for x in r]
    # xlrd used 0-indexed columns and rows in juxtaposition to xlsx rows. openpyxl uses 1-indexed columns and rows
    return [sheet.cell(x+1,ExcelCol2Index(s)+1).value for x in r]

def GetPIDs(fname,rows,column='A',sheetnum=0):
    #sheet = xlrd.open_workbook(fname).sheet_by_index(sheetnum)
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[sheetnum])
    return GetCol(sheet,column,rows)

def GetVelocityData(Dir,PIDs=None,compute_euler=True):
    ''' returns a dictionary of DataFrames from the velocity CSVs in Dir.
        If PIDs is given, tries to grab only those files, otherwise grabs all in Dir.'''
    ret = dict()
    if (PIDs == None):
        for f in glob(path.join(Dir,'*.csv')):
                ret[f] = pd.read_csv(f,delimiter=',',parse_dates=['Timestamp'],index_col=['Timestamp'])
    else:
        for i in PIDs:
            ret[i] = pd.read_csv(glob(path.join(Dir,f'{i}_*_Velocity.csv'))[0],
                delimiter=',',parse_dates=['Timestamp'],index_col=['Timestamp'])
    if(compute_euler):
        for k in ret.keys():
            for track in ('Head','LHand','RHand'):
                euler = R.from_quat(ret[k][[f'v {track} x',f'v {track} y',f'v {track} z',f'v {track} w']]).as_euler('zxy',degrees=True)
                ret[k][f'{track} Roll'],ret[k][f'{track} Pitch'],ret[k][f'{track} Yaw'] = euler[:,0],euler[:,1],euler[:,2]
    return(ret)

def GetPositionData(Dir,PIDs=None,compute_euler=True):
    ''' returns a dictionary of DataFrames from the position CSVs in Dir.
        If PIDs is given, tries to grab only those files, otherwise grabs all in Dir.'''
    ret = dict()
    if (PIDs == None):
        for f in glob(path.join(Dir,'*.csv')):
                ret[f] = pd.read_csv(f,delimiter=',',parse_dates=['Timestamp'],index_col=['Timestamp'])
    else:
        for i in PIDs:
            ret[i] = pd.read_csv(glob(path.join(Dir,f'{i}_*Motion.csv'))[0],
                delimiter=',',parse_dates=['Timestamp'],index_col=['Timestamp'])
    if(compute_euler):
        for k in ret.keys():
            for track in ('Head','Left hand','Right hand'):
                euler = R.from_quat(ret[k][[f'{track} Transform x',f'{track} Transform y',f'{track} Transform z',f'{track} Transform w']]).as_euler('zxy',degrees=True)
                ret[k][f'{track} Roll'],ret[k][f'{track} Pitch'],ret[k][f'{track} Yaw'] = euler[:,0],euler[:,1],euler[:,2]
    return(ret)

def DropData(df,todrop = ['Button','button',' w',' x',' y',' z','deltaT']):
    ''' Drops any columns from df that are named such that they contain a value that matches the given list'''
    df = df.drop([x for x in df.columns if any(l in x for l in todrop)],axis=1)
    return df

def Featurize(df,samplerate='s',aggmethods=['min','max','median','mean','std']):
    ''' Resamples and aggregates datetimeindex'd dataframes'''
    X = df.resample(samplerate).agg(aggmethods).dropna()
    X.columns = X.columns.to_flat_index()
    return X

def DivideEqualPeriods(df,n_periods=10,to_list=False):
    ''' divides a datetimeindex'd dataframe into n equal periods, returns periods in a list'''
    totaldelta = df.index.max()-df.index.min()
    subperiod = totaldelta/n_periods
    periods = [[x*subperiod+df.index.min(),(x+1)*subperiod+df.index.min()] for x in range(n_periods)]
    ret = []
    for i,(s,f) in enumerate(periods):
        res = df[s:f].copy()
        res['subphaseID'] = i
        ret.append(res)
    if to_list:
        return ret
    else:
        #this isn't efficient, but it works and was quick to write...
        return pd.concat(ret)

def FlattenWhole(dfdict,PIDs=None):
    ''' Flattens dictionary of DataFrames into single large DataFrame with order listed as PID (keys as PIDname)'''
    if PIDs == None:
        PIDs = dfdict.keys()
    for i,v in enumerate(PIDs):
        dfdict[v]['PID'] = i
        dfdict[v]['PIDname'] = v
    return pd.concat(dfdict.values())

def LeaveOneOut(df,rand_seed=None,pidcol='PID',subphasecol='subphaseID',randrange=(0,9)):
    if (rand_seed != None):
        random.seed(rand_seed)
    PIDs = df[pidcol].unique()
    leave = [random.randint(*randrange) for x in PIDs]
    leftin = pd.DataFrame()
    leftout = pd.DataFrame()
    #NB, it's easier to initialize this mask to the first value. Since we OR it with itself,
    # it shouldn't do anything weird, and it'll start at the right dimensions/dtype
    mask = (df[pidcol]==PIDs[0]) & (df[subphasecol]==leave[0])
    for pid,phase in zip(PIDs,leave):
        mask = mask | ( (df[pidcol]==pid) & (df[subphasecol]==phase) )
    leftout = df.loc[mask]
    leftin = df.loc[mask==False]
    return(leftin,leftout)

def GetOrLoadLeaveOneOuts(name="",df=None,num=20,rand_seed=None,override=False):
    if (rand_seed != None):
        random.seed(rand_seed)
    if not override and path.exists(f:=FindCache(name)):
        return pickle.load(open(f,'rb'))
    elif (df is not None):
        inoutpairs = []
        for i in range(num):
            inoutpairs.append(LeaveOneOut(df))
        pickle.dump(inoutpairs,open(FindCache(name),'wb'))
        return inoutpairs
    else:
        print("no name or dataframe provided.")


def FindCache(predFile):
    return path.join(GetCacheDir(),predFile)

def GetCacheDir():
    return(path.join(path.dirname(__file__),'LeaveOneOuts'))

def ToXy(df,Xdrop=['subphaseID','PID','PIDname'],ycol=['PID']):
    return df.drop(Xdrop,axis=1),np.ravel(df[ycol])

def ProcessToFeaturized(dfdict,todrop = ['Button','button',' w',' x',' y',' z','deltaT'],
        samplerate='s',aggmethods=['min','max','median','mean','std'],n_periods=10,to_list=False):

    prog = FloatProgress(min=0,max=len(dfdict.keys()))
    for i,k in enumerate(dfdict.keys()):
        prog.value += 1
        t = DropData(dfdict[k],todrop=todrop)
        t = Featurize(t,samplerate=samplerate,aggmethods=aggmethods)
        t = DivideEqualPeriods(t,n_periods,to_list=to_list)
        dfdict[k] = t
    return FlattenWhole(dfdict)


